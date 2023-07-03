import openpyxl
from Bio import Entrez
from datetime import datetime, timedelta
from openpyxl.utils.dataframe import dataframe_to_rows
import streamlit as st
import pandas as pd
import base64

# Set your email address for PubMed API requests
Entrez.email = "weihongtamu@gmail.com"


def search_and_save_abstracts(keyword, start_date, end_date):
    # Format the dates as required by PubMed's API
    start_date_str = start_date.strftime("%Y/%m/%d")
    end_date_str = end_date.strftime("%Y/%m/%d")

    # Search PubMed for papers within the specified time period
    search_handle = Entrez.esearch(db="pubmed", term=f'{keyword}[Title/Abstract] AND {start_date_str}:{end_date_str}[PDAT]', retmax=1000)
    search_results = Entrez.read(search_handle)
    search_handle.close()

    # Extract the list of PubMed IDs
    pmids = search_results["IdList"]

    # Fetch the records for each PubMed ID and extract the abstracts, publication date, journal, author names, and DOI
    records = []
    for pmid in pmids:
        record_handle = Entrez.efetch(db="pubmed", id=pmid, retmode="xml")
        record = Entrez.read(record_handle)["PubmedArticle"][0]
        records.append(record)
        record_handle.close()

    # Create a new workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write the headers
    sheet.cell(row=1, column=1, value="Title")
    sheet.cell(row=1, column=2, value="Abstract")
    sheet.cell(row=1, column=3, value="Publication Date")
    sheet.cell(row=1, column=4, value="Journal")
    sheet.cell(row=1, column=5, value="Authors")
    sheet.cell(row=1, column=6, value="DOI")

    # Write the article details to the spreadsheet
    for i, record in enumerate(records):
        article_title = record["MedlineCitation"]["Article"]["ArticleTitle"]
        abstract = record["MedlineCitation"]["Article"].get("Abstract", {}).get("AbstractText", [""])[0]
        publication_date = record["MedlineCitation"]["Article"]["Journal"]["JournalIssue"]["PubDate"]
        journal = record["MedlineCitation"]["Article"]["Journal"]["Title"]
        authors = [author["LastName"] + " " + author["Initials"] for author in record["MedlineCitation"]["Article"]["AuthorList"]]

        # Get the DOI if available
        doi = ""
        if "ELocationID" in record["MedlineCitation"]["Article"]:
            doi = record["MedlineCitation"]["Article"]["ELocationID"][0]

        # Convert the publication_date dictionary to a string format
        year = publication_date.get("Year", "")
        month = publication_date.get("Month", "")
        day = publication_date.get("Day", "")
        publication_date_str = f"{year} {month} {day}"

        sheet.cell(row=i + 2, column=1, value=article_title)
        sheet.cell(row=i + 2, column=2, value=abstract)
        sheet.cell(row=i + 2, column=3, value=publication_date_str)
        sheet.cell(row=i + 2, column=4, value=journal)
        sheet.cell(row=i + 2, column=5, value=", ".join(authors))
        sheet.cell(row=i + 2, column=6, value=doi)

    # Save the workbook
    output_file = f"{keyword.replace(' ', '_')}_abstracts.xlsx"
    workbook.save(output_file)

    # Convert the worksheet to a DataFrame
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)
    df = pd.DataFrame(data, columns=["Title", "Abstract", "Publication Date", "Journal", "Authors", "DOI"])

    # Sort the DataFrame by the "Publication Date" column
    df.sort_values(by="Publication Date", inplace=True)

    return df


# Streamlit App
def main():
    st.title("PubMed Abstract Search")

    # Sidebar - Date Range and Keyword Inputs
    st.sidebar.title("Search Filters")
    keyword = st.sidebar.text_input("Keyword")
    start_date = st.sidebar.date_input("Start Date")
    end_date = st.sidebar.date_input("End Date")

    if start_date and end_date:
        if start_date > end_date:
            st.error("Error: The start date must be before the end date.")
            return

        if keyword:
            # Convert the date inputs to datetime objects
            start_date = datetime.combine(start_date, datetime.min.time())
            end_date = datetime.combine(end_date, datetime.min.time())

            # Search and save abstracts
            df = search_and_save_abstracts(keyword, start_date, end_date)

            # Display the DataFrame
            st.write(df)

            # Download the output file
            output_file = df.to_csv(index=False)
            b64 = base64.b64encode(output_file.encode()).decode()  # Convert to base64 encoding
            href = f'<a href="data:file/csv;base64,{b64}" download="{keyword.replace(" ", "_")}_abstracts.csv">Download Output</a>'
            st.sidebar.markdown(href, unsafe_allow_html=True)
        else:
            st.warning("Please enter a keyword.")
    else:
        st.warning("Please select a start and end date.")


if __name__ == "__main__":
    main()
